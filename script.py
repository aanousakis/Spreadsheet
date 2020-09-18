import openpyxl

# xlsx input file name
inputFile = 'data.xlsx'

#xlsx output file name
outputFile = 'report.xlsx'

#Data starting position 
startingPos_row = '7'
startingPos_col = 'B'

#Transaction types dictionary
transactionTypes = dict()
###################################################

# replace invalid characters from sheet name
def formatSheetName(sheetName):

    s = sheetName.replace('[', '(')
    s =         s.replace(']', ')')
    s =         s.replace('/', '_')
    s =         s.replace('\\', '_')
    s =         s.replace(':', '=')
    s =         s.replace('?', '@')
    s =         s.replace('*', '@')
    
    return s


def formatData(type, s):
    #print('{}, {}'.format(type, s.split('\n')[0]))

    return s.split('\n')[0]
##########################################################

wb = openpyxl.load_workbook(inputFile)
print(wb.sheetnames)

#proccess the first sheet
workingSheet = wb['Κινήσεις Λογαριασμών']

#skip the first 7 rows
iter_row = iter(workingSheet.values)
for i in range(int(startingPos_row)-1):
    next(iter_row)


# read data from spreadsheet
for row in iter_row:
    print(row)

    #exit loop
    if row[0] is None:
        break

    if row[1] in transactionTypes:
        #append new transaction to transaction type
        transactionTypes.get(row[1]).append(row)
    else:
        #add new transaction type
        transactionTypes.update({row[1] : [row]})
   
for k, v in transactionTypes.items():
       # print(k, v)
    for q in v:
        print(q)


#output data
wb = openpyxl.Workbook() 

for k, v  in transactionTypes.items():
    #create and name new sheet
    try:
        if ws is None: # chech if ws is initialized
            pass 
        else:
            ws = wb.create_sheet(title=formatSheetName(k))
    except NameError:
        # ws does not exist at all
        ws = wb.active 
        ws.title = formatSheetName(k)

    #set width size
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 1.25 * len(k.strip())
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50


    tranDictionary = dict()

    #format transaction description and calculate sum
    for transaction in v:

        transaction = list(transaction)
        transaction[3] = formatData(k, transaction[3])
        transaction = tuple(transaction)
        ws.append(transaction)

        tranDictionary.update({transaction[3] : tranDictionary.get(transaction[3], 0) + transaction[4]})


    print("rows = {}".format(len(v)))
    #add cumulative data to sheet

    t = ("", "", "Σύνολο")
    ws.append(t)
    ws.insert_rows(len(v) + 1, 5)
    for k1, v1 in tranDictionary.items():
        print(k1, v1)
        t = ("", "", "", k1, v1)
        ws.append(t)

wb.save(outputFile)